#!/usr/bin/python
#-*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

file='C:\\Users\\Administrator\\Desktop\\hgvs_added.xlsx'
wb = load_workbook(file)
sheet = wb.active
list = []
dict = {'A':'T', 'C':'G', 'G':'C', 'T':'A'}
new_list = []

for i in range(1,sheet.max_row):
	if sheet.cell(row = i,column = 10).value == "-":
		if "non-..." not in sheet.cell(row = i, column = 14).value \
		and "null" not in sheet.cell(row = i, column = 14).value:
			list = sheet.cell(row = i,column = 14).value.split('/')
			list = [dict[x] if x in dict else x for x in list]
			new_list = '/'.join(list)
			sheet.cell(row = i,column = 14).value = new_list
			sheet.cell(row = i,column = 14).font = Font(color=colors.BLUE)
			sheet.cell(row = i,column = 10).value = "+"
			sheet.cell(row = i,column = 10).font = Font(color=colors.BLUE)
wb.save('C:\\Users\\Administrator\\Desktop\\new.xlsx')
